from openpyxl import load_workbook
from openpyxl.utils.cell import get_column_letter

import json, sys
from datetime import datetime

from pprint import pprint

source_filename = '../../data/LAT-content_Mod1_2.xlsx'
print(f"Loading {source_filename}")
wb = load_workbook(filename = source_filename)


print("Creating JSON file")
data = {}
data["steps"] = []

step_sheet = wb["Content by Screen"]

for row in range(3, step_sheet.max_row + 1):
    # if row > 8: break
    if not step_sheet[f"B{row}"].value: continue

    try:
        step = {
            "id": int(step_sheet[f"B{row}"].value) or "",
            "title": step_sheet[f"C{row}"].value or "",
            "sectionHeader": step_sheet[f"E{row}"].value or "",
            "subtitle": step_sheet[f"D{row}"].value or "",
            "type": step_sheet[f"F{row}"].value or "",
            "progressBarStage": step_sheet[f"A{row}"].value or ""
        }
    except (AttributeError, TypeError) as e:
            print(e)
            sys.exit(f"Missing button data in {row}")

    step["content"] = {}
    step["buttons"] = []
    if step["type"] == "informational":
        step["content"]["paragraphs"] = []

    for col in range(7,step_sheet.max_column+1):

        header = step_sheet[f"{get_column_letter(col)}2"].value
        val = step_sheet[f"{get_column_letter(col)}{row}"].value
        #deal with button nextId being first an sometimes empty
        if not val and header == "nextStepId" and step_sheet[f"{get_column_letter(col+4)}{row}"].value: val = "link"

        if not val: continue
        
        if type(val) == str:
            val = val.strip()
        else:
            val = int(val)

        if header == "paragraphContent":
            step["content"]["paragraphs"].append({"paragraphContent": val})
        elif header == "type" and step_sheet[f"{get_column_letter(col)}1"].value == "bullets":

            para = {"paragraphContent": {
                "type": val,
                "bullets": []
            }}

            count = 1
            while step_sheet[f"{get_column_letter(col+count)}2"].value == "bulletContent":
                if step_sheet[f"{get_column_letter(col+count)}{row}"].value:
                    para["paragraphContent"]["bullets"].append({
                                                "bulletContent": step_sheet[f"{get_column_letter(col+count)}{row}"].value.strip()
                                            })
                count += 1
            col += count 
            step["content"]["paragraphs"].append(para)
        elif header in ["src","captionText"]:
             step["content"][header] = val
        elif header in ["resetToStepId","resetText"]:
             step[header] = val
        elif header == "nextStepId":
            url = step_sheet[f"{get_column_letter(col+4)}{row}"].value 
            url = url if url else ''
            target = url.strip() if val == "link" else val
            target_title = "url" if val == "link" else "nextStepId"
            try: 
                button = {
                    "buttonText": step_sheet[f"{get_column_letter(col+1)}{row}"].value.strip(),
                    "color": step_sheet[f"{get_column_letter(col+2)}{row}"].value.strip(),
                    "textColor": step_sheet[f"{get_column_letter(col+3)}{row}"].value.strip(),
                    target_title : target   
                }
            except AttributeError as e:
                print(e)
                sys.exit(f"Missing button data in {row}")

            step["buttons"].append(button)
            col += 4








    data["steps"].append(step)

    


# pprint(data)

# FSA2002 = next((item for item in data["Forms"] if item["id"] == "FSA-2002"), None)
# pprint(FSA2002)

wb.close()


json_data = json.dumps(data, sort_keys=True, indent=4)
# print(json_data)
filename = "wizard-content.json."+datetime.strftime(datetime.now(), '%Y-%m-%d_%H:%M:%S')+"_out"
print(f"Saving output to {filename}")
with open(filename, "w") as output:
    output.write(json_data)


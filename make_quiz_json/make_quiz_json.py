from openpyxl import load_workbook
from openpyxl.utils.cell import get_column_letter

import json, sys
from datetime import datetime

from pprint import pprint

source_filename = 'LAT-quiz-content.xlsx'
print(f"Loading {source_filename}")
wb = load_workbook(filename = source_filename, data_only=True)


print("Creating JSON file")
data = {}
data["questions"] = []

step_sheet = wb["Content by Screen"]

for row in range(3, step_sheet.max_row + 1):
    # if row > 8: break
    if not step_sheet[f"B{row}"].value: continue
    try:
        step = {
            "questionId": int(step_sheet[f"A{row}"].value),
            "type": step_sheet[f"C{row}"].value or "",
            "progressBarStage": step_sheet[f"D{row}"].value or "",
            "notificationType": step_sheet[f"AS{row}"].value or "",
            "notificationTitle": step_sheet[f"AT{row}"].value or "",
            "notificationMessge": step_sheet[f"AU{row}"].value or "",
            "conditionals": str(step_sheet[f"AV{row}"].value)
        }
    except (AttributeError, TypeError) as e:
            print(e)
            sys.exit(f"Missing button data in {row}")
            
    step["content"] = {}
    step["content"]["title"] = step_sheet[f"B{row}"].value or ""
    step["content"]["paragraphs"] = []
    step["content"]["paragraphs"].append({"paragraphText": step_sheet[f"E{row}"].value or ""})
    step["content"]["paragraphs"].append({"paragraphText": step_sheet[f"F{row}"].value or ""})
    step["content"]["paragraphs"].append({"paragraphText": step_sheet[f"G{row}"].value or ""})
    step["content"]["paragraphs"].append({"paragraphText": step_sheet[f"H{row}"].value or ""})

    step["options"] = []

    option1 = {
        "optionId": int(step_sheet[f"I{row}"].value),
        "value": int(step_sheet[f"J{row}"].value),
        "nextQuestionId": int(step_sheet[f"K{row}"].value),
        "title": step_sheet[f"L{row}"].value or "",
        "url": step_sheet[f"Q{row}"].value or ""
    }
    option1["paragraphs"] = []
    option1["paragraphs"].append({"paragraphText": step_sheet[f"M{row}"].value or ""})
    option1["paragraphs"].append({"paragraphText": step_sheet[f"N{row}"].value or ""})
    option1["paragraphs"].append({"paragraphText": step_sheet[f"O{row}"].value or ""})
    option1["paragraphs"].append({"paragraphText": step_sheet[f"P{row}"].value or ""})

    step["options"].append(option1)

    option2 = {
        "optionId": int(step_sheet[f"R{row}"].value),
        "value": int(step_sheet[f"S{row}"].value),
        "nextQuestionId": int(step_sheet[f"T{row}"].value),
        "title": step_sheet[f"U{row}"].value or "",
        "url": step_sheet[f"Z{row}"].value or "",
    }
    option2["paragraphs"] = []
    option2["paragraphs"].append({"paragraphText": step_sheet[f"V{row}"].value or ""})
    option2["paragraphs"].append({"paragraphText": step_sheet[f"W{row}"].value or ""})
    option2["paragraphs"].append({"paragraphText": step_sheet[f"X{row}"].value or ""})
    option2["paragraphs"].append({"paragraphText": step_sheet[f"Y{row}"].value or ""})

    step["options"].append(option2)

    option3 = {
        "optionId": int(step_sheet[f"AA{row}"].value),
        "value": int(step_sheet[f"AB{row}"].value),
        "nextQuestionId": int(step_sheet[f"AC{row}"].value),
        "title": step_sheet[f"AD{row}"].value or "",
        "url": step_sheet[f"AI{row}"].value or ""
    }
    option3["paragraphs"] = []
    option3["paragraphs"].append({"paragraphText": step_sheet[f"AE{row}"].value or ""})
    option3["paragraphs"].append({"paragraphText": step_sheet[f"AF{row}"].value or ""})
    option3["paragraphs"].append({"paragraphText": step_sheet[f"AG{row}"].value or ""})
    option3["paragraphs"].append({"paragraphText": step_sheet[f"AH{row}"].value or ""})

    step["options"].append(option3)

    option4 = {
        "optionId": int(step_sheet[f"AJ{row}"].value),
        "value": int(step_sheet[f"AK{row}"].value),
        "nextQuestionId": int(step_sheet[f"AL{row}"].value),
        "title": step_sheet[f"AM{row}"].value or "",
        "url": step_sheet[f"AR{row}"].value or ""
    }
    option4["paragraphs"] = []
    option4["paragraphs"].append({"paragraphText": step_sheet[f"AN{row}"].value or ""})
    option4["paragraphs"].append({"paragraphText": step_sheet[f"AO{row}"].value or ""})
    option4["paragraphs"].append({"paragraphText": step_sheet[f"AP{row}"].value or ""})
    option4["paragraphs"].append({"paragraphText": step_sheet[f"AQ{row}"].value or ""})

    step["options"].append(option4)

    data["questions"].append(step)

    


# pprint(data)

# FSA2002 = next((item for item in data["Forms"] if item["id"] == "FSA-2002"), None)
# pprint(FSA2002)

wb.close()


json_data = json.dumps(data, sort_keys=True, indent=4)
# print(json_data)
filename = "quiz-content.json."
print(f"Saving output to {filename}")
with open(filename, "w") as output:
    output.write(json_data)


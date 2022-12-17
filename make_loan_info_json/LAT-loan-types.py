from openpyxl import load_workbook
from openpyxl.utils.cell import get_column_letter

import json, sys
from datetime import datetime

from pprint import pprint

source_filename = 'LAT-loan-types.xlsx'
print(f"Loading {source_filename}")
wb = load_workbook(filename = source_filename, data_only=True)


print("Creating JSON file")
data = {}
data["definitions"] = {}
data["categories"] = []
data["loanType"] = []

step_sheet = wb["definitions"]
for row in range(2, step_sheet.max_row + 1):
    # if row > 8: break
    if not step_sheet[f"B{row}"].value: continue
    try:
        definitions = {
            "maximumDollarAmount": step_sheet[f"A{row}"].value,
            "InterestRate": step_sheet[f"B{row}"].value,
            "loanTerm": step_sheet[f"C{row}"].value,
            "downPaymentRequirement": step_sheet[f"D{row}"].value
        }
    except (AttributeError, TypeError) as e:
            print(e)
            sys.exit(f"Missing button data in {row}")

data["definitions"] = definitions

step_sheet = wb["categories"]
for row in range(2, step_sheet.max_row + 1):
    # if row > 8: break
    if not step_sheet[f"B{row}"].value: continue
    try:
        categories = {
            "type": step_sheet[f"A{row}"].value,
            "description": step_sheet[f"B{row}"].value,
            "note": step_sheet[f"C{row}"].value,
            "hasMicroloan": step_sheet[f"D{row}"].value
        }
    except (AttributeError, TypeError) as e:
            print(e)
            sys.exit(f"Missing button data in {row}")
    data["categories"].append(categories)


step_sheet = wb["loanType"]
for row in range(2, step_sheet.max_row + 1):
    # if row > 8: break
    if not step_sheet[f"B{row}"].value: continue
    try:
        loanType = {
            "loanID": step_sheet[f"A{row}"].value  or "",
            "type": step_sheet[f"B{row}"].value  or "",
            "category": step_sheet[f"C{row}"].value or "",
            "formsRequired": step_sheet[f"D{row}"].value  or "",
            "formsRecommended": step_sheet[f"E{row}"].value  or "",
            "formsAdditional": step_sheet[f"F{row}"].value  or "",
            "name": step_sheet[f"G{row}"].value  or "",
            "description": step_sheet[f"H{row}"].value  or "",
            "descriptionLong": step_sheet[f"I{row}"].value  or "",
            "descriptionLongNote": step_sheet[f"J{row}"].value  or "",
            "image": step_sheet[f"K{row}"].value  or "",
            "imageAltText": step_sheet[f"L{row}"].value  or "",
            "maximumDollarAmount": step_sheet[f"M{row}"].value  or "",
            "maximumDollarAmountDescription": step_sheet[f"N{row}"].value  or "",
            "maximumDollarAmountNote": step_sheet[f"O{row}"].value or "",
            "interestRate": step_sheet[f"P{row}"].value  or "",
            "interestRateDescription": step_sheet[f"Q{row}"].value or "",
            "interestRateURL": step_sheet[f"R{row}"].value or "",
            "loanTerm": step_sheet[f"S{row}"].value  or "",
            "loanTermNote": step_sheet[f"T{row}"].value  or "",
            "loanTermDescription": step_sheet[f"U{row}"].value or "",
            "downPayment": step_sheet[f"V{row}"].value  or "",
            "downPaymentDescription": step_sheet[f"W{row}"].value or "",
            "otherRequirements": step_sheet[f"X{row}"].value  or "",
        }
    except (AttributeError, TypeError) as e:
            print(e)
            sys.exit(f"Missing button data in {row}")
    data["loanType"].append(loanType)
															
																						

# pprint(data)

wb.close()


json_data = json.dumps(data, sort_keys=True, indent=4)
# print(json_data)
filename = "LAT-loan-types.json."
print(f"Saving output to {filename}")
with open(filename, "w") as output:
    output.write(json_data)

